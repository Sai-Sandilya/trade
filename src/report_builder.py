"""
report_builder.py - Export backtest results to Excel and PDF.

Excel: multi-sheet workbook with Summary, Trade Log, Equity Curve,
       Monthly Breakdown, and Risk Metrics.
PDF:   single-page report with key metrics and a text equity summary.

Usage:
    from report_builder import build_excel_report, build_pdf_report

    xlsx_bytes = build_excel_report(summary, trade_log, equity, metrics)
    pdf_bytes  = build_pdf_report(summary, metrics, cfg_label="Balanced")

    # In Streamlit:
    st.download_button("Download Excel", xlsx_bytes, "report.xlsx", ...)
    st.download_button("Download PDF",   pdf_bytes,  "report.pdf",  ...)
"""

import io
import logging
from datetime import datetime, timezone

import pandas as pd

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def build_excel_report(
    summary:   pd.DataFrame,
    trade_log: pd.DataFrame,
    equity:    pd.DataFrame,
    metrics:   dict,
    cfg_label: str = "Backtest",
) -> bytes:
    """
    Build a multi-sheet Excel workbook and return the raw bytes.

    Sheets:
      1. Summary        — per-ticker portfolio summary
      2. Risk Metrics   — Sharpe, Sortino, CAGR, drawdown etc.
      3. Trade Log      — every trade with date, ticker, trigger, price, shares
      4. Equity Curve   — daily portfolio values
      5. Monthly P&L    — month-by-month gains/losses
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    wb = openpyxl.Workbook()

    # -- Style helpers ---------------------------------------------------------
    HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
    HEADER_FONT  = Font(color="FFFFFF", bold=True)
    TITLE_FONT   = Font(size=14, bold=True, color="1E3A5F")
    POSITIVE_FILL = PatternFill("solid", fgColor="E8F5E9")
    NEGATIVE_FILL = PatternFill("solid", fgColor="FFEBEE")
    THIN_BORDER  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    def _write_df(ws, df: pd.DataFrame, start_row: int = 1, title: str = ""):
        if title:
            ws.cell(row=start_row, column=1, value=title).font = TITLE_FONT
            start_row += 1

        df_out = df.reset_index()
        for col_idx, col_name in enumerate(df_out.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=str(col_name))
            cell.font       = HEADER_FONT
            cell.fill       = HEADER_FILL
            cell.alignment  = Alignment(horizontal="center")
            cell.border     = THIN_BORDER

        for row_idx, row in enumerate(df_out.itertuples(index=False), start_row + 1):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="right" if isinstance(val, (int, float)) else "left")

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    # -- Sheet 1: Summary ------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Summary"
    run_label = f"{cfg_label} — Generated {datetime.now(tz=timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws1["A1"] = run_label
    ws1["A1"].font = Font(size=12, italic=True, color="555555")
    ws1.merge_cells("A1:H1")

    disp = summary.copy()
    for col in ["total_invested_usd", "market_value_usd", "unrealized_pnl_usd", "total_pnl_usd"]:
        if col in disp.columns:
            disp[col] = disp[col].apply(lambda v: f"${v:,.2f}" if pd.notna(v) else "—")
    if "total_pnl_pct" in disp.columns:
        disp["total_pnl_pct"] = disp["total_pnl_pct"].apply(lambda v: f"{v:+.2f}%" if pd.notna(v) else "—")
    _write_df(ws1, disp, start_row=3, title="Portfolio Summary")

    # -- Sheet 2: Risk Metrics -------------------------------------------------
    ws2 = wb.create_sheet("Risk Metrics")
    metric_rows = [
        ("Total Return",          f"{metrics.get('total_return',     float('nan')) * 100:.2f}%"),
        ("CAGR",                  f"{metrics.get('cagr',             float('nan')) * 100:.2f}%"),
        ("Sharpe Ratio",          f"{metrics.get('sharpe',           float('nan')):.3f}"),
        ("Sortino Ratio",         f"{metrics.get('sortino',          float('nan')):.3f}"),
        ("Max Drawdown",          f"{metrics.get('max_drawdown',     float('nan')) * 100:.2f}%"),
        ("Calmar Ratio",          f"{metrics.get('calmar',           float('nan')):.3f}"),
        ("Annualised Volatility", f"{metrics.get('annualised_vol',   float('nan')) * 100:.2f}%"),
        ("Monthly Win Rate",      f"{metrics.get('win_rate_monthly', float('nan')) * 100:.1f}%"),
    ]
    _write_df(ws2, pd.DataFrame(metric_rows, columns=["Metric", "Value"]),
              start_row=1, title="Risk & Performance Metrics")

    # -- Sheet 3: Trade Log ----------------------------------------------------
    ws3 = wb.create_sheet("Trade Log")
    tl  = trade_log.copy()
    if "date" in tl.columns:
        tl["date"] = pd.to_datetime(tl["date"]).dt.strftime("%Y-%m-%d")
    _write_df(ws3, tl, start_row=1, title="Trade Log")

    # Colour BUY/SELL rows
    action_col = list(tl.reset_index().columns).index("action") + 1 if "action" in tl.columns else None
    if action_col:
        for row in ws3.iter_rows(min_row=3, max_row=ws3.max_row):
            action_cell = row[action_col - 1]
            val = str(action_cell.value or "")
            fill = POSITIVE_FILL if "BUY" in val else NEGATIVE_FILL if "SELL" in val else None
            if fill:
                for cell in row:
                    cell.fill = fill

    # -- Sheet 4: Equity Curve -------------------------------------------------
    ws4 = wb.create_sheet("Equity Curve")
    eq  = equity.copy()
    eq.index = pd.to_datetime(eq.index).strftime("%Y-%m-%d")
    eq  = eq.reset_index().rename(columns={"index": "Date"})
    _write_df(ws4, eq, start_row=1, title="Daily Portfolio Equity Curve")

    # -- Sheet 5: Monthly P&L --------------------------------------------------
    ws5 = wb.create_sheet("Monthly P&L")
    from performance_tracker import monthly_equity_breakdown
    monthly = monthly_equity_breakdown(equity)
    if not monthly.empty:
        _write_df(ws5, monthly, start_row=1, title="Month-by-Month P&L")
        # Colour positive/negative months
        change_col = list(monthly.columns).index("change_usd") + 2  # +1 for reset_index, +1 for 1-based
        for row in ws5.iter_rows(min_row=3, max_row=ws5.max_row):
            try:
                val = row[change_col - 1].value
                if isinstance(val, (int, float)):
                    fill = POSITIVE_FILL if val >= 0 else NEGATIVE_FILL
                    for cell in row:
                        cell.fill = fill
            except Exception:
                pass
    else:
        ws5["A1"] = "Not enough data for monthly breakdown (need 2+ months of equity data)"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def build_pdf_report(
    summary:   pd.DataFrame,
    metrics:   dict,
    cfg_label: str = "Backtest",
    tickers:   list[str] | None = None,
) -> bytes:
    """
    Build a clean single-page PDF report and return the raw bytes.
    Uses fpdf2 (pip install fpdf2).
    """
    try:
        from fpdf import FPDF
    except ImportError:
        raise ImportError("fpdf2 is required: pip install fpdf2")

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # -- Title -----------------------------------------------------------------
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(30, 58, 95)
    pdf.cell(0, 12, "DCA Portfolio Backtest Report", new_x="LMARGIN", new_y="NEXT", align="C")

    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(100, 100, 100)
    run_time = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    pdf.cell(0, 6, f"{cfg_label}  |  Generated {run_time}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(4)

    # -- Divider ---------------------------------------------------------------
    pdf.set_draw_color(30, 58, 95)
    pdf.set_line_width(0.5)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(4)

    # -- Portfolio Summary table -----------------------------------------------
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(30, 58, 95)
    pdf.cell(0, 8, "Portfolio Summary", new_x="LMARGIN", new_y="NEXT")

    headers = ["Ticker", "Invested", "Market Value", "Unrealised P&L", "P&L %", "Trades"]
    col_w   = [25, 35, 35, 35, 25, 20]

    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(30, 58, 95)
    pdf.set_text_color(255, 255, 255)
    for h, w in zip(headers, col_w):
        pdf.cell(w, 7, h, border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 9)
    for ticker, row in summary.iterrows():
        pnl      = float(row.get("unrealized_pnl_usd", 0))
        pnl_pct  = float(row.get("total_pnl_pct", 0))
        invested = float(row.get("total_invested_usd", 0))
        mkt_val  = float(row.get("market_value_usd", 0))
        trades   = int(row.get("num_trades", 0))

        is_pos = pnl >= 0
        pdf.set_fill_color(232, 245, 233) if is_pos else pdf.set_fill_color(255, 235, 238)
        pdf.set_text_color(0, 0, 0)

        vals = [
            str(ticker),
            f"${invested:,.0f}",
            f"${mkt_val:,.0f}",
            f"${pnl:+,.0f}",
            f"{pnl_pct:+.2f}%",
            str(trades),
        ]
        for v, w in zip(vals, col_w):
            pdf.cell(w, 6, v, border=1, align="C", fill=True)
        pdf.ln()

    # Totals row
    total_inv = float(summary["total_invested_usd"].sum())
    total_val = float(summary["market_value_usd"].sum())
    total_pnl = total_val - total_inv
    total_pct = total_pnl / total_inv * 100 if total_inv else 0
    pdf.set_fill_color(220, 220, 220)
    pdf.set_font("Helvetica", "B", 9)
    for v, w in zip(["TOTAL", f"${total_inv:,.0f}", f"${total_val:,.0f}",
                      f"${total_pnl:+,.0f}", f"{total_pct:+.2f}%", ""], col_w):
        pdf.cell(w, 7, v, border=1, align="C", fill=True)
    pdf.ln(6)

    # -- Risk Metrics ----------------------------------------------------------
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(30, 58, 95)
    pdf.cell(0, 8, "Risk & Performance Metrics", new_x="LMARGIN", new_y="NEXT")

    metric_pairs = [
        ("Total Return",    f"{metrics.get('total_return',     float('nan')) * 100:.2f}%"),
        ("CAGR",            f"{metrics.get('cagr',             float('nan')) * 100:.2f}%"),
        ("Sharpe Ratio",    f"{metrics.get('sharpe',           float('nan')):.3f}"),
        ("Sortino Ratio",   f"{metrics.get('sortino',          float('nan')):.3f}"),
        ("Max Drawdown",    f"{metrics.get('max_drawdown',     float('nan')) * 100:.2f}%"),
        ("Calmar Ratio",    f"{metrics.get('calmar',           float('nan')):.3f}"),
        ("Ann. Volatility", f"{metrics.get('annualised_vol',   float('nan')) * 100:.2f}%"),
        ("Monthly Win Rate",f"{metrics.get('win_rate_monthly', float('nan')) * 100:.1f}%"),
    ]

    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(0, 0, 0)
    col_count = 4
    box_w     = 47
    box_h     = 14

    for i, (label, value) in enumerate(metric_pairs):
        if i % col_count == 0 and i > 0:
            pdf.ln(box_h)
        is_good = _metric_is_good(label, value)
        pdf.set_fill_color(232, 245, 233) if is_good else pdf.set_fill_color(245, 245, 245)
        x = pdf.get_x()
        y = pdf.get_y()
        pdf.rect(x, y, box_w - 1, box_h, "DF")
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_xy(x + 1, y + 1)
        pdf.cell(box_w - 2, 5, label, align="C")
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_xy(x + 1, y + 6)
        pdf.cell(box_w - 2, 6, value, align="C")
        pdf.set_xy(x + box_w, y)

    pdf.ln(box_h + 6)

    # -- Disclaimer ------------------------------------------------------------
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(130, 130, 130)
    pdf.multi_cell(0, 4,
        "DISCLAIMER: This report is generated from a historical backtest simulation. "
        "Past performance does not guarantee future results. This is not financial advice. "
        "Do not make investment decisions based solely on backtested results.",
        align="C",
    )

    return bytes(pdf.output())


def _metric_is_good(label: str, value: str) -> bool:
    """Heuristic: green background if the metric value looks favourable."""
    try:
        v = float(value.replace("%", "").replace("nan", "0"))
    except ValueError:
        return False
    if "Drawdown" in label:
        return v > -20       # drawdown less than 20% is acceptable
    if "Win Rate" in label:
        return v >= 50
    return v > 0
